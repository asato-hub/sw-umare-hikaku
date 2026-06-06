from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = next((PROJECT_ROOT / "data").glob("*.xlsx"))
OUTPUT_PATH = PROJECT_ROOT / "data" / "race-stats.json"
JS_OUTPUT_PATH = PROJECT_ROOT / "data" / "race-stats.js"

ABILITIES = ["器用", "敏捷", "筋力", "生命力", "知力", "精神力"]
DICE_CONFIG = {
    "1d": {"min": 1.0, "avg": 3.5, "max": 6.0},
    "2d": {"min": 2.0, "avg": 7.0, "max": 12.0},
}


@dataclass
class RaceSource:
    row: int
    race_name: str
    source: str
    base_expressions: dict[str, Any]
    births: list[dict[str, Any]] = field(default_factory=list)


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = cell_text(value)
    if not text:
        raise ValueError("empty number")
    return float(text)


def calculate_expression(value: Any) -> dict[str, float]:
    if isinstance(value, (int, float)):
        number = float(value)
        return {"min": number, "avg": number, "max": number}

    text = cell_text(value).replace(" ", "")
    if not text:
        raise ValueError("empty expression")

    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text):
        number = float(text)
        return {"min": number, "avg": number, "max": number}

    match = re.fullmatch(r"(?:(-?\d+(?:\.\d+)?))?(1d|2d)([+-]\d+(?:\.\d+)?)?", text)
    if not match:
        raise ValueError(f"unsupported expression: {text}")

    prefix = float(match.group(1) or 0)
    dice = match.group(2)
    suffix = float(match.group(3) or 0)
    fixed = prefix + suffix
    return {
        key: fixed + DICE_CONFIG[dice][key]
        for key in ("min", "avg", "max")
    }


def add_stats(left: dict[str, float], right: dict[str, float]) -> dict[str, float]:
    return {key: left[key] + right[key] for key in ("min", "avg", "max")}


def round_number(value: float) -> int | float:
    rounded = round(value, 4)
    if rounded.is_integer():
        return int(rounded)
    return rounded


def is_base_row(row_values: list[Any]) -> bool:
    race_name = cell_text(row_values[1])
    initial_skill = cell_text(row_values[2])
    ability_values = row_values[3:9]
    return bool(race_name) and not initial_skill and all(cell_text(value) for value in ability_values)


def is_birth_row(row_values: list[Any]) -> bool:
    birth_name = cell_text(row_values[1])
    has_fixed_values = all(cell_text(row_values[index]) for index in (3, 5, 7))
    has_empty_secondary_columns = all(cell_text(row_values[index]) == "" for index in (4, 6, 8))
    return bool(birth_name) and has_fixed_values and has_empty_secondary_columns


def make_birth_stats(race: RaceSource, row: int, row_values: list[Any]) -> dict[str, Any]:
    birth_name = cell_text(row_values[1])
    initial_skill = cell_text(row_values[2])
    technique = calculate_expression(row_values[3])
    body = calculate_expression(row_values[5])
    mind = calculate_expression(row_values[7])
    source = cell_text(row_values[9])

    fixed_by_ability = {
        "器用": technique,
        "敏捷": technique,
        "筋力": body,
        "生命力": body,
        "知力": mind,
        "精神力": mind,
    }

    stats = {"min": {}, "avg": {}, "max": {}}
    for ability in ABILITIES:
        calculated = add_stats(race.base_expressions[ability], fixed_by_ability[ability])
        for kind in ("min", "avg", "max"):
            stats[kind][ability] = round_number(calculated[kind])

    return {
        "row": row,
        "birthName": birth_name,
        "initialSkill": initial_skill,
        "source": source,
        "fixed": {
            "technique": {
                "raw": cell_text(row_values[3]),
                **{key: round_number(value) for key, value in technique.items()},
            },
            "body": {
                "raw": cell_text(row_values[5]),
                **{key: round_number(value) for key, value in body.items()},
            },
            "mind": {
                "raw": cell_text(row_values[7]),
                **{key: round_number(value) for key, value in mind.items()},
            },
        },
        "stats": stats,
    }


def aggregate_race(race: RaceSource) -> dict[str, Any]:
    if not race.births:
        raise ValueError(f"{race.race_name} has no births")

    stats = {"min": {}, "avg": {}, "max": {}}
    for kind in ("min", "avg", "max"):
        for ability in ABILITIES:
            values = [birth["stats"][kind][ability] for birth in race.births]
            stats[kind][ability] = round_number(sum(values) / len(values))

    birth_average_range = {"min": {}, "mean": {}, "max": {}}
    for ability in ABILITIES:
        values = [birth["stats"]["avg"][ability] for birth in race.births]
        birth_average_range["min"][ability] = round_number(min(values))
        birth_average_range["mean"][ability] = round_number(sum(values) / len(values))
        birth_average_range["max"][ability] = round_number(max(values))

    combined_range = {"min": {}, "max": {}}
    for ability in ABILITIES:
        min_values = [birth["stats"]["min"][ability] for birth in race.births]
        max_values = [birth["stats"]["max"][ability] for birth in race.births]
        combined_range["min"][ability] = round_number(min(min_values))
        combined_range["max"][ability] = round_number(max(max_values))

    return {
        "raceName": race.race_name,
        "source": race.source,
        "excelRow": race.row,
        "birthCount": len(race.births),
        "baseExpressions": {
            ability: cell_text(race.base_expressions[f"{ability}_raw"])
            for ability in ABILITIES
        },
        "stats": stats,
        "birthAverageRange": birth_average_range,
        "combinedRange": combined_range,
        "births": race.births,
    }


def compare_summary_totals(workbook: Any, aggregated_races: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if len(workbook.worksheets) < 3:
        return []

    summary_sheet = workbook.worksheets[2]
    races_by_name = {race["raceName"]: race for race in aggregated_races}
    warnings = []

    for row_number in range(4, summary_sheet.max_row + 1):
        race_name = cell_text(summary_sheet.cell(row_number, 2).value)
        expected_total = summary_sheet.cell(row_number, 5).value
        if not race_name or expected_total in (None, "") or race_name not in races_by_name:
            continue

        race = races_by_name[race_name]
        calculated_total = round(sum(race["stats"]["avg"].values()), 2)
        difference = round(calculated_total - float(expected_total), 2)
        if abs(difference) > 0.01:
            warnings.append(
                {
                    "raceName": race_name,
                    "summarySheetTotal": round_number(float(expected_total)),
                    "calculatedAverageTotal": round_number(calculated_total),
                    "difference": round_number(difference),
                    "note": "Detailed six-ability values were used for JSON aggregation.",
                }
            )

    return warnings


def main() -> None:
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    worksheet = workbook.worksheets[0]

    races: list[RaceSource] = []
    active_races: list[RaceSource] = []
    sheet_names = [sheet.title for sheet in workbook.worksheets]

    for row_number in range(4, worksheet.max_row + 1):
        row_values = [worksheet.cell(row_number, col).value for col in range(1, 12)]

        if all(cell_text(value) == "" for value in row_values):
            active_races = []
            continue

        if is_base_row(row_values):
            base_raw = dict(zip(ABILITIES, row_values[3:9], strict=True))
            base_expressions: dict[str, Any] = {}
            for ability, raw_value in base_raw.items():
                base_expressions[ability] = calculate_expression(raw_value)
                base_expressions[f"{ability}_raw"] = raw_value

            race = RaceSource(
                row=row_number,
                race_name=cell_text(row_values[1]),
                source=cell_text(row_values[9]),
                base_expressions=base_expressions,
            )
            races.append(race)
            active_races.append(race)
            continue

        if is_birth_row(row_values):
            if not active_races:
                raise ValueError(f"birth row without active race at row {row_number}")
            for race in active_races:
                race.births.append(make_birth_stats(race, row_number, row_values))

    aggregated_races = [aggregate_race(race) for race in races]
    validation_warnings = compare_summary_totals(workbook, aggregated_races)

    output = {
        "meta": {
            "diceConfig": DICE_CONFIG,
        },
        "metadata": {
            "sourceFile": "data/生まれ表一覧.xlsx",
            "sourceSheet": worksheet.title,
            "availableSheets": sheet_names,
            "abilities": ABILITIES,
            "ruleDiceConfig": DICE_CONFIG,
            "displayDiceConfig": DICE_CONFIG,
            "aggregationRule": "Each birth belonging to the same race is averaged with equal weight.",
            "notes": [
                "Race rows contain the six racial dice expressions.",
                "Birth rows contain technique/body/mind fixed values; each fixed value is added to two abilities.",
                "Rows such as タビット（パイカ種） and ウィークリング（タンノズ） are treated as separate race variants because their racial dice expressions differ.",
            ],
            "validationWarnings": validation_warnings,
        },
        "races": aggregated_races,
    }

    json_text = json.dumps(output, ensure_ascii=False, indent=2)
    OUTPUT_PATH.write_text(json_text + "\n", encoding="utf-8")
    JS_OUTPUT_PATH.write_text(f"window.RACE_STATS = {json_text};\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")
    print(f"Wrote {JS_OUTPUT_PATH}")
    print(f"Races: {len(output['races'])}")


if __name__ == "__main__":
    main()
